from django.shortcuts import redirect, render
from .models import Producto,Categoria,Umedida,Proveedor,Cliente,DetalleCotizacion,Cotizacion,Tipo_Cotizacion,Empresa,EstadoCotizacion,Venta,DetalleVenta,Tipo_Venta,Compras,DetalleCompras
from django.http import JsonResponse,HttpResponse
from django.views.generic import  View

from django.db.models import Q,Max
from django.contrib.auth.models import User
from django.db import transaction

from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.auth import authenticate,logout,login as auth_login
from django.urls import reverse
from django.contrib.auth.decorators import login_required
import pandas as pd
from openpyxl import load_workbook


def Login(request):
    template_name = "Login.html"
    response_data = {}
    context = {}
    empresas=Empresa.objects.filter(state=True)
    if request.method == 'GET':
        context={
            "empresas":empresas
        }
        pass
    else:    
        if request.POST.get('action') == 'post':
            username = request.POST.get('usuario')
            password = request.POST.get('clave')
            empresa=request.POST.get('empresa')
            if len(empresa)>0:            
                nombre_empresa=Empresa.objects.get(id=empresa)
                user = authenticate(username=username, password=password)
                if user:
                    auth_login(request, user)
                    response_data['loggedin'] = True
                    response_data['username'] = user.first_name
                    response_data['completo'] = '{} {}'.format(user.first_name, user.last_name)
                    response_data['id'] = user.id
                    response_data['url'] = reverse('index_home')
                    request.session['usuario'] = '{} {}'.format(user.first_name, user.last_name)
                    request.session['loggedin'] = True
                    request.session['idempresa'] = empresa
                    request.session['empresa'] = nombre_empresa.nombre
                    request.session['logoempresa'] =nombre_empresa.logo_path
                    request.session['logoempresacotizacion'] =nombre_empresa.logo_path_cotizacion

                    response_data['msg'] = 'Bienvenido es un placer tenerlo por aca !!!'
                else:
                    response_data['loggedin'] = False
                    request.session['loggedin'] = False
                    response_data['msg'] = "Nombre de usuario o contraseña incorrecto!!!"
            else:
                 response_data['loggedin'] = False
                 request.session['loggedin'] = False
                 response_data['msg'] = "Seleccione una Empresa!!!"        
            return JsonResponse(response_data)
    return render(request, template_name,context)    

def Logoutapp(request):
    logout(request)
    return redirect('login')

def handler404(request, exception):
    context = {}
    response = render(request, "404.html", context=context)
    response.status_code = 404
    return response


def handler500(request):
    context = {}
    response = render(request, "500.html", context=context)
    response.status_code = 500
    return response

@login_required
def index(request):
    template_name="Index.html"
    return render(request,template_name)

#region Productos
@login_required
def Productos_listado(request):
    #SELECT * FROM PRODUCTOS WHERE STATE=TRUE
    productos = Producto.objects.filter(state=True)
    context = {
        "productos": productos,
    }
    template_name = "Listado_productos.html"
    return render(request, template_name, context)

@login_required
def Productos_visualizar(request,id):
    template_name = "Visualizar_productos.html"
    producto=Producto.objects.get(id=id)
    categoria=Categoria.objects.filter(state=True)
    proveedor=Proveedor.objects.filter(state=True)
    umedida=Umedida.objects.filter(state=True)
    
    #Seleccionado el color para mostrarla en el Combo
    categorias=Categoria.objects.get(id=producto.categoria.id)
    proveedores=Proveedor.objects.get(id=producto.proveedor.id)
    umedidas=Umedida.objects.get(id=producto.umedida.id)
    
    context = {"productos":producto,
               "Categorias":categoria,
               "Proveedores":proveedor,
               "Umedida":umedida,
               "categoriaseleccionado":categorias.nombre,
               "proveedorseleccionado":proveedores.nombre,
               "umedidaseleccionado":umedidas.nombre
               }
    return render(request, template_name,context)   

@login_required
def UpdateCrudProductos(request):
    response_data = {}
    template_name="Visualizar_productos.html"
    if request.POST.get('action') =='actualizar_productos':
        try:
            id=request.POST.get('id')
            productos = Producto.objects.get(id=id)
            productos.nombre=request.POST.get('txtnomprod')

            categoria_producto = request.POST.get('cmbcategoria')
            
            if categoria_producto=="":
                categoria=productos.categoria
            else:
                categoria=Categoria.objects.get(id=categoria_producto)
            productos.categoria=categoria

            productos.descripcion=request.POST.get('txtdescripcion')

            proveedor_producto = request.POST.get('cmbproveedores')
            if proveedor_producto =="":
                provee=productos.proveedor
            else:    
                provee=Proveedor.objects.get(id=proveedor_producto)
            productos.proveedor=provee

            productos.costo_real=float(request.POST.get('txtcostoreal'))
            productos.costo_ofrecido=float(request.POST.get('txtcostofrecido'))
            productos.ganancia=0

            umedida_producto = request.POST.get('cmbumedida')
            if umedida_producto =="":
                umedida=productos.umedida
            else:    
                umedida=Umedida.objects.get(id=umedida_producto)
            productos.umedida=umedida

            campos_a_validar = ['txtcostomayoreo', 'txtstockmin', 'txtstockmax','txtstock']
            valores, errores = validar_campos(request, campos_a_validar)

            productos.codigo_interno=request.POST.get('txtcodinterno')
            productos.costo_mayoreo=valores.get('txtcostomayoreo', 0)
            productos.stock = valores.get('txtstock', 0)
            productos.stock_maximo = valores.get('txtstockmax', 0)
            productos.stock_minimo = valores.get('txtstockmin', 0)
            productos.ubicacion=request.POST.get('txtubicacion')
            productos.codigo_barra=request.POST.get('txtcodbarra')
            if request.FILES.get('imagen'):
                productos.imagen = request.FILES['imagen']
            productos.save()

            # if errores:
            #     return JsonResponse({'errores': errores}, status=400)
            # else:
                
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No Se Modifico el Productos Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Modifico con exito el Producto'  
            
        return JsonResponse(response_data)   
    return render(request,template_name) 

def validar_campos(request, campos):
    valores_validos = {}
    errores = {}

    for campo in campos:
        valor = request.POST.get(campo)
        
        if not valor or not valor.strip().isdigit():
            valores_validos[campo] = 0  # Asignar valor por defecto si está vacío o no es numérico
            errores[campo] = f"El campo {campo} es obligatorio y debe ser un número válido."
        else:
            valores_validos[campo] = int(valor)  # Asignar el valor convertido a entero
    
    return valores_validos, errores


class DeleteCrudProductos(View):
    def get(self, request):
        id1 = request.GET.get('id', None)
        if not id1:
            return JsonResponse({'error': 'ID no proporcionado.'}, status=400)
        try:
            Producto.objects.get(id=id1).delete()
            data = {'deleted': True}
            return JsonResponse(data)
        except Producto.DoesNotExist:
            return JsonResponse({'error': 'El elemento no existe.'}, status=404)

    def post(self, request):
        id1 = request.POST.get('id', None)
        if not id1:
            return JsonResponse({'error': 'ID no proporcionado.'}, status=400)
         
        try:
            Producto.objects.get(id=id1).delete()
            data = {'deleted': True}
            return JsonResponse(data)
        except Producto.DoesNotExist:
            return JsonResponse({'error': 'El elemento no existe.'}, status=404)
        


@login_required    
def Productos_form(request):
    template_name = "Nuevo_productos.html"
    categoria=Categoria.objects.filter(state=True)
    proveedor=Proveedor.objects.filter(state=True)
    umedida=Umedida.objects.filter(state=True)
    context = {
        "Categorias":categoria,"Proveedores":proveedor,"Umedida":umedida
    }
    return render(request, template_name,context)   

@login_required
def CreateCrudProductos(request):
    response_data = {}
    template_name="Nuevo_productos.html"
    if request.POST.get('action') =='registrar_productos':
        try:
            campos_a_validar = ['txtcostomayoreo', 'txtstockmin', 'txtstockmax','txtstock']
            valores, errores = validar_campos(request, campos_a_validar)
            codigo_interno=request.POST.get('txtcodinterno')
            codigo_barra=request.POST.get('txtcodbarra')
            costo_mayoreo=valores.get('txtcostomayoreo', 0)
            stock=valores.get('txtstock', 0)
            stock_maximo = valores.get('txtstockmax', 0)
            stock_minimo = valores.get('txtstockmin', 0)
            ubicacion=request.POST.get('txtubicacion')
            nombre_producto = request.POST.get('txtnomprod')
            categoria_producto = request.POST.get('cmbcategoria')
            des_producto = request.POST.get('txtdescripcion')
            proveedor_producto = request.POST.get('cmbproveedores')
            costoreal = request.POST.get('txtcostoreal')
            costofrecido = request.POST.get('txtcostofrecido')
            ganancia =0
            unidadmedida_producto = request.POST.get('cmbumedida')
            categoria=Categoria.objects.get(id=categoria_producto)
            proveedor=Proveedor.objects.get(id=proveedor_producto)
            umedida=Umedida.objects.get(id=unidadmedida_producto)
            imagen = request.FILES.get('imagen', None)
            Producto.objects.create(nombre=nombre_producto,
                                      categoria=categoria,
                                      descripcion=des_producto,
                                      proveedor=proveedor,
                                      costo_real=costoreal,
                                      costo_ofrecido=costofrecido,
                                      ganancia=ganancia,
                                      umedida=umedida,
                                      codigo_interno= codigo_interno,
                                      codigo_barra=codigo_barra,
                                      costo_mayoreo=costo_mayoreo,
                                      stock=stock,
                                      stock_maximo=stock_maximo,
                                      stock_minimo=stock_minimo,
                                      ubicacion=ubicacion,
                                      se_importo=False,
                                      imagen=imagen
                                      )
            
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No se Registro el Producto Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Registro con exito el Producto'   
        return JsonResponse(response_data)   
    return render(request,template_name) 


class AutocompleteServicios(View):
    def get(self, request, *args, **kwargs):
        query = self.request.GET.get('query', '')
        items = Producto.objects.filter(Q(nombre__icontains=query))
        #suggestions = [item.nombre for item in items]
        suggestions = [{'nombre': item.nombre, 'precio': item.costo_real,'precio_venta':item.costo_ofrecido,'stock':item.stock,'id': item.pk } for item in items]
        return JsonResponse({'suggestions': suggestions})
    


@login_required
def importar_excel_productos(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        try:
            # Lee el archivo Excel
            df = pd.read_excel(excel_file)
            # Reemplaza valores NaN con cadenas vacías o valores por defecto
            df.fillna('', inplace=True)
            # Convierte los datos a una lista
            data = df.values.tolist()
            return JsonResponse({'data': data})
        except Exception as e:
            # Maneja cualquier error al leer el archivo
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'No file uploaded'}, status=400)
 


@login_required
def cargar_productos_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            archivo = request.FILES['file']
            
            if not archivo.name.endswith('.xlsx'):
                return JsonResponse({'error': 'El archivo debe ser un archivo Excel (.xlsx)'}, status=400)

            workbook = load_workbook(archivo)
            sheet = workbook.active

            errores = []  # Lista para almacenar errores

            with transaction.atomic():  # Inicia una transacción atómica
                for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) != 12:
                        errores.append(f"Error en la fila {index}: se esperaban 12 columnas, pero se encontraron {len(row)}")
                        continue
                    try:
                        # Desempaquetar fila
                        Codigo, Nombre, PreciodeCosto, PreciodeVenta, PreciodeMayoreo, Unidaddemedida, Stock, StockMinimo, StockMaximo, CategoriaV, ProveedorV, Ubicacion = row

                        # Validar datos obligatorios
                        if not Nombre or PreciodeCosto in [None, ""] or PreciodeVenta in [None, ""] or Stock in [None, ""]:
                            errores.append(f"Faltan datos obligatorios en la fila {index}")
                            continue

                        # Validar y convertir valores numéricos
                        def validar_numero(valor):
                            if valor in ["No aplica", None, ""]:
                                return 0
                            try:
                                return float(valor)
                            except ValueError:
                                raise ValueError(f"Valor numérico inválido en la fila {index}")

                        costo_real = validar_numero(PreciodeCosto)
                        costo_ofrecido = validar_numero(PreciodeVenta)
                        costo_mayoreo = validar_numero(PreciodeMayoreo)
                        stock = validar_numero(Stock)
                        stock_minimo = validar_numero(StockMinimo)
                        stock_maximo = validar_numero(StockMaximo)

                        # Obtener objetos relacionados
                        ObjCategoria = get_or_default(Categoria, CategoriaV, "SIN CATEGORIA")
                        ObjProveedor = get_or_default(Proveedor, ProveedorV, "SIN PROVEEDOR")
                        ObjUmedida = get_or_default(Umedida, Unidaddemedida, "SIN UMEDIDA")

                        # Crear producto
                        Producto.objects.create(
                            codigo_interno=Codigo,
                            nombre=Nombre,
                            costo_real=costo_real,
                            costo_ofrecido=costo_ofrecido,
                            costo_mayoreo=costo_mayoreo,
                            stock=int(stock),
                            stock_minimo=int(stock_minimo),
                            stock_maximo=int(stock_maximo),
                            ubicacion=Ubicacion,
                            categoria=ObjCategoria,
                            proveedor=ObjProveedor,
                            umedida=ObjUmedida,
                            se_importo=True
                        )
                    except Exception as e:
                        errores.append(f"Error en la fila {index}: {str(e)}")

                # Si hay errores, se cancela la transacción
                if errores:
                    raise ValueError("Errores encontrados en el archivo Excel")

            # Si todo va bien, se retorna éxito
            return JsonResponse({'mensaje': 'Todos los productos se cargaron correctamente'})

        except Exception as e:
            return JsonResponse({'message': 'No se realizó la carga debido a errores', 'detalles': errores}, status=400)

    return JsonResponse({'error': 'No se recibió un archivo válido'}, status=400)

def get_or_default(model, field_name, default_value):
    """
    Intenta obtener un objeto del modelo por el campo 'field_name'.
    Si no se encuentra, devuelve el objeto correspondiente al valor por defecto.
    """
    if not field_name:  # Si el valor es vacío o None
        return model.objects.get(nombre=default_value)
    
    # Intentar obtener el objeto por nombre
    obj = model.objects.filter(nombre=field_name).first()
    
    # Si no existe, devolver el objeto por defecto
    return obj if obj else model.objects.get(nombre=default_value) 
#endregion Productos

#region Categorias
@login_required
def Categorias_listado(request):
    categorias = Categoria.objects.filter(state=True)
    context = {
        "Categorias": categorias,
    }
    template_name = "Listado_categorias.html"
    return render(request, template_name, context)

@login_required
def Categoria_form(request):
    template_name = "Nueva_categoria.html"
    return render(request, template_name)

@login_required
def CreateCrudCategorias(request):
    response_data = {}
    template_name="Nueva_categoria.html"
    if request.POST.get('action') =='registrar_categoria':
        try:
            nombre_categoria = request.POST.get('txtnomcat')
            Categoria.objects.create(nombre=nombre_categoria)
            
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No se Registro la Categoria Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Registro con exito la Categoria'   
        return JsonResponse(response_data)   
    return render(request,template_name) 

@login_required
def Categoria_visualizar(request,id):
    template_name = "Visualizar_categoria.html"
    categoria=Categoria.objects.get(id=id)
    context = {"categorias":
               categoria}
    return render(request, template_name,context)

@login_required
def UpdateCrudCategorias(request):
    response_data = {}
    template_name="Visualizar_categoria.html"
    if request.POST.get('action') =='actualizar_categoria':
        try:
            id=request.POST.get('id')
            categoria = Categoria.objects.get(id=id)
            categoria.nombre=request.POST.get('txtnomcat')
            categoria.save()
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No Se Modifico la Categoria Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Modifico con exito la Categoria'  
            
        return JsonResponse(response_data)   
    return render(request,template_name) 


class DeleteCrudCategorias(View):
    def  get(self, request):
        id1 = request.GET.get('id', None)
        Categoria.objects.get(id=id1).delete()
        data = {
            'deleted': True
        }
        return JsonResponse(data)
#endregion categorias    
    
#region Proveedores    
@login_required
def Proveedores_listado(request):
    proveedores = Proveedor.objects.filter(state=True)
    context = {
        "Proveedores": proveedores,
    }
    template_name = "Listado_proveedores.html"
    return render(request, template_name, context)    

@login_required
def Proveedor_form(request):
    template_name = "Nuevo_proveedor.html"
    return render(request, template_name)

@login_required
def CreateCrudProveedores(request):
    response_data = {}
    template_name="Nuevo_proveedor.html"
    if request.POST.get('action') =='registrar_proveedor':
        try:
            nombre_proveedor = request.POST.get('txtnomprov')
            ruc_proveedor = request.POST.get('txtrucprov')
            dir_proveedor = request.POST.get('txtdirprov')
            tlf_proveedor = request.POST.get('txttelprov')
            email_proveedor = request.POST.get('txtemailprov')
            contacto_proveedor = request.POST.get('txtcontactoprov')
            
            Proveedor.objects.create(nombre=nombre_proveedor,direccion=dir_proveedor,ruc=ruc_proveedor,telefono=tlf_proveedor,email=email_proveedor,contacto=contacto_proveedor)
            
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No se Registro El Proveedor Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Registro con exito El Proveedor'   
        return JsonResponse(response_data)   
    return render(request,template_name) 


class DeleteCrudProveedores(View):
    def  get(self, request):
        id1 = request.GET.get('id', None)
        Proveedor.objects.get(id=id1).delete()
        data = {
            'deleted': True
        }
        return JsonResponse(data)

@login_required    
def Proveedor_visualizar(request,id):
    template_name = "Visualizar_proveedor.html"
    proveedor=Proveedor.objects.get(id=id)
    context = {"proveedor":proveedor
               }
    return render(request, template_name,context)   

@login_required
def UpdateCrudProveedor(request):
    response_data = {}
    template_name="Visualizar_proveedor.html"
    if request.POST.get('action') =='actualizar_proveedor':
        try:
            id=request.POST.get('id')
            proveedor = Proveedor.objects.get(id=id)
            proveedor.nombre=request.POST.get('txtnom')
            proveedor.ruc=request.POST.get('txtruc')
            proveedor.direccion=request.POST.get('txtdir')
            proveedor.telefono=request.POST.get('txtfono')
            proveedor.email=request.POST.get('txtemail')
            proveedor.contacto=request.POST.get('txtcontacto')
            proveedor.save()
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No Se Modifico el Proveedor Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Modifico con exito El Proveedor'  
            
             
        return JsonResponse(response_data)   
    return render(request,template_name)  
#endregion Proveedores   

#region Clientes
@login_required
def Clientes_listado(request):
    clientes = Cliente.objects.filter(state=True)
    context = {
        "Clientes": clientes,
    }
    template_name = "Listado_clientes.html"
    return render(request, template_name, context)   

@login_required
def Cliente_form(request):
    template_name = "Nuevo_cliente.html"
    return render(request, template_name)

@login_required
def CreateCrudClientes(request):
    response_data = {}
    template_name="Nuevo_cliente.html"
    if request.POST.get('action') =='registrar_cliente':
        try:
            nombre_cliente = request.POST.get('txtnomclie')
            ruc_cliente = request.POST.get('txtrucclie')
            dir_cliente = request.POST.get('txtdirclie')
            tlf_cliente = request.POST.get('txttelclie')
            email_cliente = request.POST.get('txtemailclie')
            contacto_cliente = request.POST.get('txtcontactoclie')
            
            Cliente.objects.create(nombre=nombre_cliente,direccion=dir_cliente,ruc=ruc_cliente,telefono=tlf_cliente,email=email_cliente,contacto=contacto_cliente)
            
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No se Registro El Cliente Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Registro con exito El Cliente'   
        return JsonResponse(response_data)   
    return render(request,template_name) 


class DeleteCrudClientes(View):
    def  get(self, request):
        id1 = request.GET.get('id', None)
        Cliente.objects.get(id=id1).delete()
        data = {
            'deleted': True
        }
        return JsonResponse(data)

@login_required    
def Cliente_visualizar(request,id):
    template_name = "Visualizar_cliente.html"
    cliente=Cliente.objects.get(id=id)
    context = {"cliente":cliente
               }
    return render(request, template_name,context)   

@login_required
def UpdateCrudCliente(request):
    response_data = {}
    template_name="Visualizar_cliente.html"
    if request.POST.get('action') =='actualizar_cliente':
        try:
            id=request.POST.get('id')
            cliente = Cliente.objects.get(id=id)
            cliente.nombre=request.POST.get('txtnom')
            cliente.ruc=request.POST.get('txtruc')
            cliente.direccion=request.POST.get('txtdir')
            cliente.telefono=request.POST.get('txtfono')
            cliente.email=request.POST.get('txtemail')
            cliente.contacto=request.POST.get('txtcontacto')
            cliente.save()
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No Se Modifico el Cliente Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Modifico con exito El Cliente'  
            
             
        return JsonResponse(response_data)   
    return render(request,template_name)      

class AutocompleteClientes(View):
    def get(self, request, *args, **kwargs):
        query = self.request.GET.get('query', '')
        items = Cliente.objects.filter(Q(nombre__icontains=query))
        #suggestions = [item.nombre for item in items]
        suggestions = [{'nombre': item.nombre, 'ruc': item.ruc,'contacto':item.contacto,'id':item.pk} for item in items]
        return JsonResponse({'suggestions': suggestions})
    
#endregion Clientes   
    
@login_required
def Listar_cotizaciones(request):
    template_name="Listado_cotizaciones.html"
    cotizaciones=Cotizacion.objects.filter(state=True,convertido_factura=False)#orm QUERYSET
    cotizaciones_ordenadas = cotizaciones.order_by('-fecha_cotizacion')
    context = {
        "cotizaciones": cotizaciones_ordenadas,
    }
    return render(request,template_name,context)    

@login_required
def incrementar_detalle(request):
    # Inicializamos el contador en 0 si es la primera vez que se accede
    if 'contador' not in request.session:
        request.session['contador'] = 0
    # Incrementamos el contador en 1
    lstdetallecotizacion = request.session.get('lstdetallecotizacion', [])    
    request.session['contador'] =int(len(lstdetallecotizacion)+1)
    contador=request.session['contador']
    return contador

#ACA HAGO DOS COSAS ACTUALIZAR Y GRABAR EN LAS LISTAS EN MEMORIA
@login_required
def Grabar_item_cotizacion(request):
    response_data = {}
    lstdetallecotizacion = request.session.get('lstdetallecotizacion', [])
    template_name="cotizacion.html"
    if request.POST.get('action') =='registrar_cotizacion':
        try:
           
            iddetalle=request.POST.get('iddetalle')
            idservicio=request.POST.get('id_servicio')
            cotizacion=request.POST.get('numcotizacion')
            detalle = request.POST.get('detalle')
            precio = request.POST.get('precio')
            preciocompra = request.POST.get('preciocompra')
            cantidad = request.POST.get('cantidad')
            total = request.POST.get('total')
            costocompra= request.POST.get('costocompra')
           
            if len(iddetalle)!=0:
                for indice in lstdetallecotizacion:
                    
                    if int(iddetalle)==int(indice["iddetalle"]):
                        
                        if len(idservicio)!=0:
                            objservicio=Producto.objects.get(pk=idservicio)
                            servicio = objservicio.nombre
                            id=idservicio
                            lstdetallecotizacion[int(iddetalle)-1]["idservicio"]=idservicio
                            lstdetallecotizacion[int(iddetalle)-1]["servicio"]=servicio
                        else:    
                            lstdetallecotizacion[int(iddetalle)-1]["detalle"]=detalle
                            lstdetallecotizacion[int(iddetalle)-1]["precio"]=precio
                            lstdetallecotizacion[int(iddetalle)-1]["preciocompra"]=preciocompra
                            lstdetallecotizacion[int(iddetalle)-1]["cantidad"]=cantidad
                          
                            lstdetallecotizacion[int(iddetalle)-1]["costo"]=total
                            lstdetallecotizacion[int(iddetalle)-1]["costocompra"]=costocompra
                        request.session['lstdetallecotizacion'] = lstdetallecotizacion
                        break
            else:
                
                objservicio=Producto.objects.get(pk=idservicio)
                servicio = objservicio.nombre
                id=idservicio
                item_detalle_cotizacion={
                            "iddetalle":int(incrementar_detalle(request)),
                            "cotizacion":cotizacion,
                            "idservicio":id,
                            "servicio":servicio,
                            "detalle":detalle,
                            "precio":precio,
                            "preciocompra":preciocompra,
                            "cantidad":cantidad,
                            
                            "costo":total,
                            "costocompra":costocompra
                        }
                lstdetallecotizacion.append(item_detalle_cotizacion)
                # Guardar el diccionario en la sesión
                request.session['lstdetallecotizacion'] = lstdetallecotizacion
                
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No se Registro el Servicio Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Registro con exito el Servicio'  
        return JsonResponse(response_data)   
    return render(request,template_name) 

#ACA CREO LA COTIZACION EN LA BD CABECERA Y DETALLE
@login_required
def Cotizacion_crear(request,numcot,tipo,actualizar):
    response_data = {}
    context={}
    template_name="cotizacion.html"
    if request.method == 'GET':
        lstdetallecotizacion = request.session.get('lstdetallecotizacion', {})
        context = {
             "detalle": lstdetallecotizacion,
        }
    else:    
        if request.POST.get('action') =='registrar_cotizacion_cabecera':
            try:
                opciones=request.POST.get('opciones')
                if opciones=="1":
                    objtipo_cotizacion = Tipo_Cotizacion.objects.get(pk=1)
                else:
                    objtipo_cotizacion = Tipo_Cotizacion.objects.get(pk=2)
                nombre_evento = request.POST.get('evento')
                capacidad = request.POST.get('capacidad')
                lugar = request.POST.get('lugar')    
                numcotizacion= request.POST.get('numcotizacion')    
                idcliente = request.POST.get('idcliente')
                objcliente=Cliente.objects.get(id=idcliente)
                fecha_cotizacion = request.POST.get('fecha')
                comentario = request.POST.get('comentario')
                totalFee = 0
                mas_sin_igv=True
                obj_user=User.objects.get(pk=1)
                persona_creo_cotiza=obj_user
                total=request.POST.get('totalSubt')
                totalcompra=0
                obj_estado_cot=EstadoCotizacion.objects.get(pk=1)
                estado=obj_estado_cot
                codempresa=int(request.session.get('idempresa'))
                obj_empresa=Empresa.objects.get(pk=codempresa)
                con_que_empresa=obj_empresa
                lstdetallecotizacion = request.session.get('lstdetallecotizacion', [])
                
                item=grabar_cotizacion(lstdetallecotizacion, numcotizacion,objtipo_cotizacion,objcliente,nombre_evento,fecha_cotizacion,capacidad,lugar,comentario,totalFee,mas_sin_igv,persona_creo_cotiza,total,totalcompra,estado,con_que_empresa)
               
                if not item:
                    raise ValueError("Debe seleccionar por lo menos un Servicio")
                elif item!=True:
                    raise ValueError("Debe Ingresar todos los campos Correctamente")
            except ValueError as error:
                response_data['flag'] = False
                response_data['msg'] = f'Debe Ingresar todos los campos Correctamente {error} '       
            except Exception as error:
                response_data['flag'] = False
                response_data['msg'] = f'Debe Ingresar todos los campos Correctamente {error} '       
            else:
                response_data['flag'] = True
                response_data['msg'] = 'Se Registro con exito la Cotizacion'
                request.session['lstdetallecotizacion'] = []   
            return JsonResponse(response_data)   
       
    return render(request, template_name, context)



def grabar_cotizacion(data_items, numcot,tipocot,cliente,evento,fecha,capacidad,lugar,comentario,fee,mas_sin_igv,persona_creo_cotiza,total,totalcompra,estado,con_que_empresa):
    detalle_agregado = False  # Bandera para indicar si se ha agregado algún detalle
    try:
        # Comenzar una transacción
        with transaction.atomic():
            # Si hay registros en el diccionario, guardar los detalles de la cotización
            if data_items:
                 # Guardar el número de cotización en la cabecera
                Cotizacion.objects.create(nrocotizacion=numcot,
                        tipo_cotizacion=tipocot,
                        cliente=cliente,
                        nombre_evento=evento,
                        fecha_cotizacion=fecha,
                        capacidad=capacidad,
                        lugar=lugar,
                        comentario=comentario,
                        fee=fee,
                        mas_sin_igv=mas_sin_igv,
                        persona_creo_cotiza=persona_creo_cotiza,
                        total=total,
                        totalcompra=totalcompra,
                        estado=estado,
                        con_que_empresa=con_que_empresa)
                # Obtener la instancia de Cotizacion correspondiente a numcot
                cotizacion = Cotizacion.objects.get(pk=numcot)
                for item in data_items:
                    objservicio = Producto.objects.get(pk=item["idservicio"])
                    DetalleCotizacion.objects.create(
                        cotizacion=cotizacion,
                        servicio=objservicio,
                        detalle=item["detalle"],
                        precio=item["precio"],
                        preciocompra=item["preciocompra"],
                        cantidad=item["cantidad"],
                       
                        costo=item["costo"],
                        costocompra=item["costocompra"]
                    )
                detalle_agregado = True  # Se ha agregado al menos un detalle    
        return detalle_agregado 
    except Exception as e:
        # Manejar errores de transacción
        print(f"Error al guardar la cotización: {e}")
        return f"Error al guardar la cotización: {e}"


@login_required
def Cancelar_cotizacion(request):
    template_name="cotizacion.html"
    if request.method == 'POST':
        if 'lstdetallecotizacion' in request.session:
            del request.session['lstdetallecotizacion']
            del request.session['contador']
            return redirect('listar_cotizaciones') 
    return render(request, template_name)    

@login_required
def Cotizacion_visualizar(request,numcot,tipo,actualizar):
    response_data = {}
    context={}
    template_name="cotizacion.html"
    if request.method == 'GET':
        
        cabeceracotizacion=Cotizacion.objects.get(pk=numcot)
        estado_cotizacion = cabeceracotizacion.estado.id
        #print(f"este es el estado de la coti {estado_cotizacion}")
        tipo_cot = actualizar
        
        lstdetallecotizacion = DetalleCotizacion.objects.filter(cotizacion=numcot,
        state=True)
        estadocotizacion=EstadoCotizacion.objects.all()

        context = {
             "detalle": lstdetallecotizacion,
             "cabecera":cabeceracotizacion,
             "estadocotizacion":estadocotizacion,
             'valor_bd': estado_cotizacion,
             "tipo_cot":tipo_cot
        }
       
       
    else:    
        if request.POST.get('action') =='registrar_cotizacion_cabecera':
            try:
                opciones=request.POST.get('opciones')
               
                if opciones=="1":
                    objtipo_cotizacion = Tipo_Cotizacion.objects.get(pk=1)
                else:
                    objtipo_cotizacion = Tipo_Cotizacion.objects.get(pk=2)
                nombre_evento = request.POST.get('evento')
                capacidad = request.POST.get('capacidad')
                lugar = request.POST.get('lugar')    
                numcotizacion= request.POST.get('numcotizacion')    
                idcliente = request.POST.get('idcliente')
                objcliente=Cliente.objects.get(id=idcliente)
                fecha_cotizacion = request.POST.get('fecha')
                comentario = request.POST.get('comentario')
                totalFee = 0
                mas_sin_igv=True
                user_logeado=response_data['id']
                obj_user=User.objects.get(pk=user_logeado)
                persona_creo_cotiza=obj_user
                total=request.POST.get('totalSubt')
                obj_estado_cot=EstadoCotizacion.objects.get(pk=1)
                estado=obj_estado_cot
                codempresa=int(request.session.get('idempresa'))
                obj_empresa=Empresa.objects.get(pk=codempresa)
                con_que_empresa=obj_empresa
                lstdetallecotizacion = request.session.get('lstdetallecotizacion', [])
               
                item=grabar_cotizacion(lstdetallecotizacion, numcotizacion,objtipo_cotizacion,objcliente,nombre_evento,fecha_cotizacion,capacidad,lugar,comentario,totalFee,mas_sin_igv,persona_creo_cotiza,total,estado,con_que_empresa)
                if not item:
                    raise ValueError("Debe seleccionar por lo menos un Servicio")
                elif item!=True:
                    raise ValueError("Debe Ingresar todos los campos Correctamente")
            except ValueError as error:
                response_data['flag'] = False
                response_data['msg'] = f'Debe Ingresar todos los campos Correctamente'       
            except Exception as error:
                response_data['flag'] = False
                response_data['msg'] = f'Debe Ingresar todos los campos Correctamente'       
            else:
                response_data['flag'] = True
                response_data['msg'] = 'Se Registro con exito la Cotizacion'
                request.session['lstdetallecotizacion'] = []   
            return JsonResponse(response_data)   
       
    return render(request, template_name, context)


class Eliminar_detalle_cotizacion(View):
    def  get(self, request):
        id1 = request.GET.get('id', None)
        if 'lstdetallecotizacion' in request.session:
            # Obtén la lista de la sesión
            lstdetallecotizacion = request.session['lstdetallecotizacion']
            # Verifica si se proporcionó un índice válido en la solicitud POST
            indice_eliminar = int(id1) - 1
            # Elimina el elemento de la lista por su índice
            del lstdetallecotizacion[indice_eliminar]
            # Reasignar los índices de la lista
            #lstdetallecotizacionnueva = [elemento for elemento in lstdetallecotizacion]
            for index, detalle in enumerate(lstdetallecotizacion):
                detalle['iddetalle'] = index + 1
            # Actualiza la lista en la sesión
            request.session['lstdetallecotizacion'] = lstdetallecotizacion
            data = {
                'deleted': True
            }
            print("esta es la nueva lista regenerada")
            print(lstdetallecotizacion)
        return JsonResponse(data)
    
#ACA MODIFICO Y INSERTO UNA COTIZACION CUANDO YA ESTA GRABADO EN LA BD
@login_required
def modificar_item_cotizacion(request):
    response_data = {}
    template_name="cotizacion.html"
    if request.POST.get('action') =='actualizar_item_cotizacion':
        try:
            id=request.POST.get('iddetalle')
            if len(id)==0:
                 objservicio = Producto.objects.get(pk=request.POST.get('id_servicio'))
                 numcotizacion=request.POST.get('numcotizacion')
                 objcotizacion=Cotizacion.objects.get(pk=numcotizacion)
                 DetalleCotizacion.objects.create(
                        cotizacion=objcotizacion,
                        servicio=objservicio,
                        detalle=request.POST.get('detalle'),
                        precio=request.POST.get('precio'),
                        preciocompra=request.POST.get('preciocompra'),
                        cantidad=request.POST.get('cantidad'),
                       
                        costo=request.POST.get('total'),
                        costocompra=request.POST.get('costocompra')
                    )
            else: 
                detalle_cotizacion = DetalleCotizacion.objects.get(pk=id)
                servicio_cambiar=request.POST.get('id_servicio')
                if len(servicio_cambiar)!=0:
                    servicio=Producto.objects.get(pk=servicio_cambiar)
                    detalle_cotizacion.servicio=servicio
                detalle_cotizacion.detalle=request.POST.get('detalle')
                detalle_cotizacion.precio=request.POST.get('precio')
                detalle_cotizacion.preciocompra=request.POST.get('preciocompra')
                detalle_cotizacion.cantidad=request.POST.get('cantidad')
                
                detalle_cotizacion.costo=request.POST.get('total')
                detalle_cotizacion.costocompra=request.POST.get('costocompra')
                detalle_cotizacion.save()
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'No Se Modifico el Servicio Correctamente {error}'
        else:
            response_data['flag'] = True
            response_data['msg'] = 'Se Modifico con exito El Servicio'  
        return JsonResponse(response_data)   
    return render(request,template_name)          
    

class EliminarItemCotizacionBD(View):
    def  get(self, request):
        id1 = request.GET.get('id', None)
        DetalleCotizacion.objects.get(id=id1).delete()
        data = {
            'deleted': True
        }
        return JsonResponse(data)
    

class EliminarCotizacionBD(View):
    def  get(self, request):
        id1 = request.GET.get('id', None)
        Cotizacion.objects.filter(nrocotizacion=id1).update(state=False)
        data = {
            'deleted': True
        }
        return JsonResponse(data)    

@login_required
def Actualizar_totales_bd(request):
    if request.method == 'POST' and request.POST.get('action') == 'actualizar_totales':
        try:
            numcotizacion= request.POST.get('numcotizacion')    
            idcliente = request.POST.get('idcliente')
            nombre_evento = request.POST.get('evento')
            capacidad = request.POST.get('capacidad')
            lugar = request.POST.get('lugar')    
            fecha_cotizacion = request.POST.get('fecha')
            comentario = request.POST.get('comentario')
            totalFee = 0
            totalSubt = request.POST.get('totalSubt')
            totalcompra = 0
            estadocotizacion=request.POST.get('estadocotizacion')
            print(f"estadodecotizacionact {estadocotizacion}")

            coti = Cotizacion.objects.get(nrocotizacion=numcotizacion)
            if len(idcliente)>0:
                objcliente=Cliente.objects.get(id=idcliente)
                coti.cliente=objcliente
            coti.nombre_evento=nombre_evento
            coti.fecha_cotizacion=fecha_cotizacion
            coti.capacidad=capacidad
            coti.lugar=lugar
            coti.comentario=comentario
            coti.fee= totalFee
            coti.total = totalSubt
            coti.totalcompra = totalcompra
            objestadocot=EstadoCotizacion.objects.get(pk=estadocotizacion)
            coti.estado=objestadocot
            coti.save()
            response_data = {'flag': True, 'msg': 'Se actualizó con éxito'}
        except Exception as e:
            response_data = {'flag': False, 'msg': f'Error: {str(e)}'}
    else:
        response_data = {'flag': False, 'msg': 'Solicitud no válida'}
    return JsonResponse(response_data)


def obtener_la_ultima_cotizacion(request):
    try:
        # Obtenemos el último objeto Cotizacion ordenado por 'nrocotizacion' de manera descendente
        ultima_cotizacion = Cotizacion.objects.order_by('-nrocotizacion').first()
        if ultima_cotizacion:
            max_valor = ultima_cotizacion.nrocotizacion
        else:
            max_valor = '10000000'
    except Cotizacion.DoesNotExist:
        max_valor = '10000000'

    return JsonResponse({'max_valor': max_valor})


@login_required
def visualizar_cotizacion(request,numcot):
     template_path="cotizacion_print.html"
     template = get_template(template_path)
     cabeceracotizacion=Cotizacion.objects.get(pk=numcot)
     lstdetallecotizacion = DetalleCotizacion.objects.filter(cotizacion=numcot,
        state=True)
     subtotal=cabeceracotizacion.total
     fee=cabeceracotizacion.fee
     total=float(subtotal) + float(fee)
     igv=float(total)*0.18
     total_ven=total+igv
     
     context = {
             "cabecera":cabeceracotizacion,
             "detalle": lstdetallecotizacion,
             "total":round(total,2),
             "igv" :round(igv,2),
             "total_ven":round(total_ven,2),
             "logopath":request.session.get('logoempresacotizacion')
        }
     #return render(request,template_path,context)

     
     html = template.render(context)
     response = HttpResponse(content_type='application/pdf')
     nombre_pdf=f"{numcot}.pdf"
     response['Content-Disposition'] = 'attachment; filename='+nombre_pdf

     pisa_status = pisa.CreatePDF(
        html, dest=response, encoding='utf-8')

     if pisa_status.err:
        return HttpResponse('Error al generar PDF: %s' % html)

     return response
    
@login_required
def Listar_ventas(request):
    template_name="Listado_Ventas.html"
    ventas=Venta.objects.filter(state=True)#orm QUERYSET
    ventas_ordenadas = ventas.order_by('fecha_venta')
    context = {
        "ventas": ventas_ordenadas,
    }
    return render(request,template_name,context) 
    

@login_required
def Cotizacion_a_Ventas(request,numcot):
    response_data = {}
    context={}
    template_name="Ventas.html"
    #Generando el codigo de la venta
    ultima_venta = Venta.objects.order_by('-nroventa').first()
    if ultima_venta:
        max_valor = ultima_venta.nroventa
    else:
        max_valor = '10000000'
    # Incrementar el valor como un número entero
    nuevo_valor = str(int(max_valor) + 1)    
    cabeceracotizacion=Cotizacion.objects.get(pk=numcot)
    lstdetallecotizacion = DetalleCotizacion.objects.filter(cotizacion=numcot,state=True)
    tipo_venta=Tipo_Venta.objects.filter(state=True)
    if request.method == 'GET':
        context = {
             "detalle": lstdetallecotizacion,
             "cabecera":cabeceracotizacion,
             "tipoventa":tipo_venta
        }
    else:   
            if request.POST.get('action') == 'cotizacion_a_ventas': 
                try:
                    valortp=request.POST.get('cmbtipoventa')
                    tipoventa=Tipo_Venta.objects.get(pk=valortp)
                    venta_cabecera=Venta.objects.create(
                        nroventa=nuevo_valor,
                        nrodocumento=request.POST.get('numdocumento'),
                        nrocotizacion=numcot,
                        cliente=cabeceracotizacion.cliente,
                        fecha_venta=request.POST.get('fecha'),
                        comentario=request.POST.get('comentario'),
                        persona_creo_venta=cabeceracotizacion.persona_creo_cotiza,
                        total_neto=cabeceracotizacion.total,
                        total_igv=request.POST.get('totalIgv'),
                        total_descuento=request.POST.get('descuento'),
                        tipo_venta=tipoventa,
                        con_que_empresa=cabeceracotizacion.con_que_empresa
                    )
                    #Guardando el detalle de la venta
                    #   
                    for item in lstdetallecotizacion:
                        objproducto = Producto.objects.get(pk=item.servicio.pk)
                        DetalleVenta.objects.create(
                            venta=venta_cabecera,
                            producto=objproducto,
                            detalle=item.detalle,
                            cantidad=item.cantidad
                        )
                          # Restar la cantidad vendida del stock del producto
                        objproducto.stock -= item.cantidad
                        objproducto.save()  # Guardar los cambios en la base de datos
                    cabeceracotizacion.convertido_factura=True
                    cabeceracotizacion.save(update_fields=['convertido_factura'])
                except Exception as error:
                    response_data['flag'] = False
                    response_data['msg'] = f'No se Registro La Venta Correctamente {error}'        
                else:  
                    response_data['flag'] = True  
                    response_data['msg'] = 'Se Registro con exito La Venta' 
            return JsonResponse(response_data) 
    return render(request, template_name, context)

@login_required
def Visualizar_Venta(request,nroventa):
    context={}
    template_name="Visualizar_Venta.html"
    cabeceraventa=Venta.objects.get(pk=nroventa)
    lstdetalleventa = DetalleVenta.objects.filter(venta=nroventa)
    tipo_venta=Tipo_Venta.objects.filter(state=True)
    #Seleccionado el color para mostrarla en el Combo
    tipoventa=Tipo_Venta.objects.get(id=cabeceraventa.tipo_venta.id)
    context = {
                "detalle": lstdetalleventa,
                "cabecera":cabeceraventa,
                "tipoventa":tipo_venta,
                "tipoventaseleccionada":tipoventa
        }
    return render(request, template_name, context)


#LO REFERENTE A COMPRAS

@login_required
def Listar_compras(request):
    template_name="Listado_Compras.html"
    compras=Compras.objects.filter(state=True)#orm QUERYSET
    compras_ordenadas = compras.order_by('-fecha')
    context = {
        "compras": compras_ordenadas,
    }
    return render(request,template_name,context)    


#GUARDA LOS ITEMS DE LA COMPRA EN MEMORIA
@login_required
def Grabar_item_compra(request):
    response_data = {}
    template_name="compras.html"
    lstdetallecompra = request.session.get('lstdetallecompra', [])
    compra=request.POST.get('numcompra')
    idservicio=request.POST.get('id_servicio')
    precioventa = request.POST.get('precioventa')
    preciocompra = request.POST.get('preciocompra')
    cantidad = request.POST.get('cantidad')
    subtotal = float(preciocompra) * int(cantidad)
    
    try:
        if request.POST.get('action') =='registrar_compras':
               
                objservicio=Producto.objects.get(pk=idservicio)
                servicio = objservicio.nombre
                id=idservicio
                item_detalle_compra={
                                "iddetalle":int(incrementar_detalle_compras(request)),
                                "compra":compra,
                                "idservicio":id,
                                "servicio":servicio,
                                "precioventa":precioventa,
                                "preciocompra":preciocompra,
                                "cantidad":cantidad,
                                "subtotal":subtotal,
                            }
                lstdetallecompra.append(item_detalle_compra)
                    # Guardar el diccionario en la sesión
                request.session['lstdetallecompra'] = lstdetallecompra
                
            
        else:
                
                iddetalle = request.POST.get('iddetalle')
                for indice in lstdetallecompra:
                    if int(iddetalle)==int(indice["iddetalle"]):
                        if len(idservicio)!=0:
                            objservicio=Producto.objects.get(pk=idservicio)
                            servicio = objservicio.nombre
                            id=idservicio
                            lstdetallecompra[int(iddetalle)-1]["idservicio"]=idservicio
                            lstdetallecompra[int(iddetalle)-1]["servicio"]=servicio
                        else:
                            
                            lstdetallecompra[int(iddetalle)-1]["precioventa"]=precioventa
                            lstdetallecompra[int(iddetalle)-1]["preciocompra"]=preciocompra
                            lstdetallecompra[int(iddetalle)-1]["cantidad"]=cantidad
                            lstdetallecompra[int(iddetalle)-1]["subtotal"]=subtotal
                            request.session['lstdetallecompra'] = lstdetallecompra
                            break
    except Exception as error:
        response_data['flag'] = False
        response_data['msg'] = f'No se Registro el Producto Correctamente {error}'
    else:
        response_data['flag'] = True
        response_data['msg'] = 'Se Registro con exito el Producto'  
        return JsonResponse(response_data)   
    return render(request,template_name) 


#GUARDA LA CABECERA Y EL DETALLE DE LA COMPRA
@login_required
def Compra_crear(request):
    request.session['opcion'] = 'create'
    response_data = {}
    #request.session['lstdetallecompra'] = [] #Esta linea de codigo limpia la sesion de memoria
    context={}
    template_name="compras.html"
    proveedor=Proveedor.objects.filter(state=True)
    lstdetallecompra = request.session.get('lstdetallecompra', {})
    #Generando el codigo de compra
    ultima_compra = Compras.objects.order_by('-numero_compra').first()
    if ultima_compra:
        max_valor = ultima_compra.numero_compra
    else:
        max_valor = '10000000'
    # Incrementar el valor como un número entero
    nuevo_valor = str(int(max_valor) + 1) 
    if request.method == 'GET':
        context = {
            "Proveedores":proveedor,
            "detallecompra":lstdetallecompra,
            "idcompra":nuevo_valor
        }
        
    else:    
        if request.POST.get('action') == 'grabar_compra': 
                try:
                    
                    obj_user=User.objects.get(pk=1)
                    persona_creo_cotiza=obj_user    
                    codempresa=int(request.session.get('idempresa'))
                    obj_empresa=Empresa.objects.get(pk=codempresa)
                    valorprov=request.POST.get('cmbproveedores')
                    numdocumento=request.POST.get('numdocumento')
                    proveedor=Proveedor.objects.get(pk=valorprov)
                    compra_cabecera=Compras.objects.create(
                        numero_compra=nuevo_valor,
                        numero_documento=numdocumento,
                        proveedor=proveedor,
                        fecha=request.POST.get('fecha'),
                        detalle=request.POST.get('comentario'),
                        persona_creo_venta=persona_creo_cotiza,
                        total=request.POST.get('totalGran'),
                        con_que_empresa=obj_empresa
                    )
                    #Guardando el detalle de la Compra
                    for item in lstdetallecompra:
                        objproducto = Producto.objects.get(pk=item["idservicio"])
                        DetalleCompras.objects.create(
                            compra=compra_cabecera,
                            producto=objproducto,
                            precio_compra=item["preciocompra"],
                            precio_venta=item["precioventa"],
                            cantidad=item["cantidad"],
                            subtotal=item["subtotal"],
                        )
                          # Aumenta la cantidad de stock producto
                        objproducto.stock += int(item["cantidad"])
                        objproducto.save()  # Guardar los cambios en la base de datos

                        if 'lstdetallecompra' in request.session:
                            del request.session['lstdetallecompra']
                            
                    
                except Exception as error:
                    response_data['flag'] = False
                    response_data['msg'] = f'No se Registro La Compra Correctamente {error}'        
                else:  
                    response_data['flag'] = True  
                    response_data['msg'] = 'Se Registro con exito La Compra' 
        return JsonResponse(response_data) 
    return render(request, template_name, context)


@login_required
def incrementar_detalle_compras(request):
    # Inicializamos el contador en 0 si es la primera vez que se accede
    if 'contadorcom' not in request.session:
        request.session['contadorcom'] = 0
    # Incrementamos el contador en 1
    lstdetallecompra = request.session.get('lstdetallecompra', [])    
    request.session['contadorcom'] =int(len(lstdetallecompra)+1)
    contador=request.session['contadorcom']
    return contador


class Eliminar_detalle_compra(View):
    def post(self, request):
            id1 = request.POST.get('id', None)
            try:
                if not id1:
                    return JsonResponse({'deleted': False, 'error': 'ID no proporcionado.'}, status=400)

                id1 = int(id1)
                

                if 'lstdetallecompra' in request.session:
                    lstdetallecompra = request.session['lstdetallecompra']

                    # Verificar si la lista no está vacía
                    if not lstdetallecompra:
                        return JsonResponse({'deleted': False, 'error': 'La lista ya está vacía.'}, status=400)

                    # Verificar si el índice es válido
                    if id1 <= 0 or id1 > len(lstdetallecompra):
                        return JsonResponse({'deleted': False, 'error': 'Índice fuera de rango.'}, status=400)

                    # Eliminar el elemento
                    del lstdetallecompra[id1 - 1]

                    # Reasignar índices si aún hay elementos en la lista
                    if lstdetallecompra:
                        for index, detalle in enumerate(lstdetallecompra):
                            detalle['iddetalle'] = index + 1

                    # Actualizar la sesión
                    request.session['lstdetallecompra'] = lstdetallecompra
                    

                    return JsonResponse({'deleted': True})
                else:
                    return JsonResponse({'deleted': False, 'error': 'Lista no encontrada en la sesión.'}, status=400)
            except ValueError:
                return JsonResponse({'deleted': False, 'error': 'ID inválido.'}, status=400)
            except Exception as e:
                print("Error:", str(e))
                return JsonResponse({'deleted': False, 'error': str(e)}, status=500)
         

@login_required
def Visualizar_Compra(request,nrocompra):
    request.session['opcion'] = 'update'
    context={}
    template_name="Visualizar_Compra.html"
    proveedor=Proveedor.objects.filter(state=True)
    
    cabeceracompra=Compras.objects.get(numero_compra=nrocompra)
    lstdetallecompra = DetalleCompras.objects.filter(compra=cabeceracompra)
    context = {
        "Proveedores":proveedor,
        "detalle": lstdetallecompra,
        "cabecera":cabeceracompra       
          }  
    return render(request, template_name, context)

class EliminarItemComprasBD(View):
    def get(self, request):
        id1 = request.GET.get('id', None)
        if not id1:
            return JsonResponse({'error': 'ID no proporcionado.'}, status=400)
        try:
            DetalleCompras.objects.get(id=id1).delete()
            data = {'deleted': True}
            return JsonResponse(data)
        except DetalleCompras.DoesNotExist:
            return JsonResponse({'error': 'El elemento no existe.'}, status=404)

    def post(self, request):
        id1 = request.POST.get('id', None)
        if not id1:
            return JsonResponse({'error': 'ID no proporcionado.'}, status=400)
         
        try:
            #Elimino el Stock de ese Producto   
            objdetcom=DetalleCompras.objects.get(id=id1)
            idproducto=objdetcom.producto.id
            objproducto = Producto.objects.get(pk=idproducto)   
            objproducto.stock -= objdetcom.cantidad
            objproducto.save()
            #Luego elimino el producto de la BD y me deberia devolver el Stock
            DetalleCompras.objects.get(id=id1).delete()
            data = {'deleted': True}
            return JsonResponse(data)
        except DetalleCompras.DoesNotExist:
            return JsonResponse({'error': 'El elemento no existe.'}, status=404)
        

#GUARDA LOS ITEMS DE LA COMPRA EN LA BD CUANDFO SE ACTUALIZA UNA COMPRA
@login_required
def Grabar_item_compraBD(request):
    response_data = {}
    if request.method == 'POST':
        try:
            compra = request.POST.get('numcompra')
            iddetalle=request.POST.get('iddetalle')
            idservicio = request.POST.get('id_servicio', '').strip() or 'DEFAULT_ID_SERVICIO'
            precioventa = request.POST.get('precioventa')
            preciocompra = request.POST.get('preciocompra')
            cantidad = request.POST.get('cantidad')

            if not compra or not idservicio or not precioventa or not preciocompra or not cantidad:
                raise ValueError("Faltan datos obligatorios.")

            preciocompra = float(preciocompra)
            cantidad = int(cantidad)
            subtotal = preciocompra * cantidad

            if request.POST.get('action') == 'registrar_comprasBD':
                objproducto = Producto.objects.get(pk=idservicio)
                objcompra = Compras.objects.get(numero_compra=compra)
                
                DetalleCompras.objects.create(
                    compra=objcompra,
                    producto=objproducto,
                    precio_compra=preciocompra,
                    precio_venta=precioventa,
                    cantidad=cantidad,
                    subtotal=subtotal
                )
                # Aumentar el stock del producto
                objproducto.stock += cantidad
                objproducto.save()

                response_data['flag'] = True
                response_data['msg'] = 'Se registró con éxito el producto.'

            elif request.POST.get('action') == 'actualizar_comprasBD':
                objproducto = Producto.objects.get(pk=idservicio)
                detalle = DetalleCompras.objects.filter(pk=iddetalle).first()
                if not detalle:
                    raise Exception("El detalle con el id proporcionado no existe.")
                
                # Primera Logica.- Cuando se actualiza un item de compra
                # Resto lo que se agrego anteriormente y lo grabo
                cantidad_anterior=detalle.cantidad
                print(f"este es la cantidad anterior {cantidad_anterior}")
                objproducto.stock -= cantidad_anterior
                objproducto.save()

                detalle.precio_compra = preciocompra
                detalle.precio_venta = precioventa
                detalle.cantidad = cantidad
                subtotal = float(preciocompra) * int(cantidad)
                detalle.subtotal = subtotal
                detalle.save()  # Guardar cambios

                # Segunda Logica.- Cuando se actualiza un item de compra
                # Agrego el nuevo valor que se modifico
                
                objproducto.stock += cantidad
                objproducto.save()
                print(f"este es la cantidad nueva que se va agregar {cantidad}")

                response_data = {'flag': True, 'msg': 'Detalle actualizado correctamente.'}
                

        except Producto.DoesNotExist:
            response_data['flag'] = False
            response_data['msg'] = f'El producto con ID {idservicio} no existe.'
        except Compras.DoesNotExist:
            response_data['flag'] = False
            response_data['msg'] = f'La compra con ID {compra} no existe.'
        except ValueError as ve:
            response_data['flag'] = False
            response_data['msg'] = str(ve)
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'Error: {error}'
    else:
        response_data['flag'] = False
        response_data['msg'] = 'Método no permitido.'

    return JsonResponse(response_data)


@login_required
def actualizar_cabecera_compraBD(request):
    response_data = {}
    if request.method == 'POST':
        try:
            valorprov=request.POST.get('cmbproveedores')
            proveedor=Proveedor.objects.get(pk=valorprov)
            numcompra=request.POST.get('numcompra')
            numdocumento=request.POST.get('numdocumento')
            fecha=request.POST.get('fecha'),
            detalle=request.POST.get('comentario'),
            total=request.POST.get('totalGran'),
            if isinstance(fecha, tuple):  # Si es una tupla
                fecha = fecha[0]  # Toma el primer elemento
            if request.POST.get('action') == 'registrar_comprascabeceraBD':
                compra = Compras.objects.filter(numero_compra=numcompra).first()
                if not detalle:
                    raise Exception("El detalle con el id proporcionado no existe.")
                compra.fecha = fecha
                compra.proveedor = proveedor
                if isinstance(detalle, tuple):  # Si es una tupla, desempaquétalo
                    detalle = detalle[0]
                compra.detalle = detalle
                if isinstance(total, tuple):  # Si es una tupla, desempaquétalo
                    total = total[0]
                compra.total = float(total)
                compra.numero_documento=numdocumento
                compra.save()  # Guardar cambios

                response_data = {'flag': True, 'msg': 'Compra actualizada correctamente.'}
        except Exception as error:
            response_data['flag'] = False
            response_data['msg'] = f'Error: {error}'
    else:
        response_data['flag'] = False
        response_data['msg'] = 'Método no permitido.'

    return JsonResponse(response_data)

class EliminarCompraBD(View):
    def actualizar_estado_compra(self, numero_compra):
        """
        Marca la compra como inactiva y regresa el stock de los productos asociados.
        """
        try:
            with transaction.atomic():
                # Actualiza el estado de la compra
                Compras.objects.filter(numero_compra=numero_compra).update(state=False)

                # Actualiza el stock de los productos asociados
                detalles = DetalleCompras.objects.filter(compra__numero_compra=numero_compra)
                for detalle in detalles:
                    producto = detalle.producto
                    producto.stock -= detalle.cantidad
                    producto.save()
        except Exception as e:
            return str(e)
        return None

    def post(self, request):
        numero_compra = request.GET.get('id')
        if not numero_compra:
            return JsonResponse({'error': 'El ID de la compra no fue proporcionado.'}, status=400)

        error = self.actualizar_estado_compra(numero_compra)
        if error:
            return JsonResponse({'error': f'Ocurrió un error: {error}'}, status=500)

        return JsonResponse({'deleted': True})

    def get(self, request):
        return self.post(request)  # Delegar lógica al método POST para evitar duplicación

